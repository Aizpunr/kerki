"""Parse a Kerki cup from BepInEx LiveLeaderboardLogger.log → JSON + (optional) xlsx write.

Usage:
    python parse_kerki_from_log.py 35
    python parse_kerki_from_log.py 35 --mappers 76561198125435428,76561198083300332
    python parse_kerki_from_log.py 35 --mappers 76561198050792757,76561198083300332 --write-xlsx
    python parse_kerki_from_log.py 35 --date 2026-05-10 --log "C:/path/to/LiveLeaderboardLogger.log"

Rules (kept aligned with project_kerki.md):
- Warmup: first played round on each of the 3 unique maps (3 warmup rounds).
- Aborted rounds (final LB has 0 finishers) are dropped.
- Points: 1=100, 2=80, 3=70, 4-5=60, 6-9=50, 10-16=40, 17-24=35, 25+=30
- Finalist threshold: 750 points (capped). Mappers excluded from championship contention
  (they take points off the pool by occupying slots).
- Winner = finalist who wins a round AS finalist. Up to 5. Order = round won asc.
- Finalists section sorted by qualifying-round asc.
- Last-round-rule: if a player qualifies in the LAST scoring round, no finalist tag.
- Participation filter: must be rostered ≥6 scoring rounds AND have ≥1 finish.
"""
import argparse
import json
import os
import re
import shutil
import sys
from collections import defaultdict
from datetime import date as _date

# ── Config ─────────────────────────────────────────────────────────
DEFAULT_LOG = r"C:\Program Files (x86)\Steam\steamapps\common\Zeepkist\BepInEx\LiveLeaderboardLogger.log"
PROJECT_DIR = os.path.dirname(os.path.abspath(__file__))
FINALIST_THRESHOLD = 750
MIN_ROUNDS_PLAYED = 6

POINTS_TABLE = [
    (1, 1, 100), (2, 2, 80), (3, 3, 70), (4, 5, 60),
    (6, 9, 50), (10, 16, 40), (17, 24, 35), (25, 9999, 30),
]

def points_for_pos(pos):
    for lo, hi, pts in POINTS_TABLE:
        if lo <= pos <= hi:
            return pts
    return 30

# Tab + column for a given kerki id
# Tab "Kerki 31-35" → col_offsets [1,7,13,19,25] for kerkis 31,32,33,34,35
# Tab "Kerki 36-40" → col_offsets [1,7,13,19,25] for 36,37,38,39,40
def xlsx_target(kerki_id):
    if 1 <= kerki_id <= 5:    return ("Kerki Comp Results.xlsx", "Kerki 1-5", kerki_id - 1)
    if 6 <= kerki_id <= 10:   return ("Kerki Comp Results.xlsx", "Kerki 6-10", kerki_id - 6)
    if 11 <= kerki_id <= 15:  return ("Kerki Comp Results.xlsx", "Kerki 11-15", kerki_id - 11)
    if 16 <= kerki_id <= 20:  return ("Kerki Comp Results 16-30.xlsx", "Kerki 16-20", kerki_id - 16)
    if 21 <= kerki_id <= 25:  return ("Kerki Comp Results 16-30.xlsx", "Kerki 21-25", kerki_id - 21)
    if 26 <= kerki_id <= 30:  return ("Kerki Comp Results 16-30.xlsx", "Kerki 26-30", kerki_id - 26)
    # 31+ block — extend tab list as new ones are created
    if 31 <= kerki_id <= 35:  return ("Kerki Comp Results 31+.xlsx", "Kerki 31-35", kerki_id - 31)
    if 36 <= kerki_id <= 40:  return ("Kerki Comp Results 31+.xlsx", "Kerki 36-40", kerki_id - 36)
    raise ValueError(f"No xlsx mapping for Kerki #{kerki_id} — add a row to xlsx_target()")

OFFSETS = [1, 7, 13, 19, 25]  # 5 kerkis per tab, 6 cols each (3 data + 3 spacer)

# ── Log parsing ────────────────────────────────────────────────────
def parse_log(log_path, target_kerki, restrict_date=None):
    """Return rounds + roster dict for the given kerki id (across all sessions on `restrict_date`).
    `restrict_date`: 'YYYY-MM-DD' or None. If None, takes latest session containing the kerki."""
    map_marker = f"Kerki #{target_kerki}"
    roster = {}        # sid -> name (latest seen)
    rounds = []
    current_round = None
    current_lb = None
    current_session = 0
    session_active = False

    with open(log_path, encoding="utf-8") as f:
        for line in f:
            m = re.match(r"^([\d\-T:.]+) \[LiveLeaderboardLogger\] (.+)$", line.rstrip("\n"))
            if not m:
                continue
            ts, evt = m.group(1), m.group(2)
            if restrict_date and not ts.startswith(restrict_date):
                continue
            parts = evt.split("|")
            kind = parts[0]

            if kind == "SESSION_START":
                current_session += 1
                session_active = True
                if current_round is not None:
                    current_round["last_lb"] = current_lb
                    rounds.append(current_round)
                    current_round = None
                    current_lb = None
            elif kind == "SESSION_END":
                if current_round is not None:
                    current_round["last_lb"] = current_lb
                    rounds.append(current_round)
                    current_round = None
                    current_lb = None
                session_active = False
            elif not session_active:
                continue
            elif kind == "ROSTER":
                sid, name = parts[1], parts[2]
                roster[sid] = name
                if current_round is not None:
                    current_round["roster"].add(sid)
            elif kind == "ROUND_STARTED":
                if current_round is not None:
                    current_round["last_lb"] = current_lb
                    rounds.append(current_round)
                rmap = parts[3] if len(parts) > 3 else ""
                current_round = {
                    "num": int(parts[1]), "hash": parts[2], "map": rmap,
                    "session": current_session, "last_lb": None, "ended": False,
                    "roster": set(), "ts": ts,
                }
                current_lb = None
            elif kind == "ROUND_ENDED":
                if current_round is not None:
                    current_round["ended"] = True
            elif kind == "LEADERBOARD":
                if current_round is None or current_round.get("ended"):
                    continue
                payload = parts[3] if len(parts) > 3 else ""
                entries = []
                if payload:
                    for tok in payload.split(","):
                        p = tok.split(":")
                        if len(p) >= 3:
                            entries.append({"pos": int(p[0]), "sid": p[1], "time": p[2]})
                if current_lb is None or len(entries) >= len(current_lb):
                    current_lb = entries

    if current_round is not None:
        current_round["last_lb"] = current_lb
        rounds.append(current_round)

    kerki_rounds = [r for r in rounds if map_marker in r["map"]]
    return kerki_rounds, roster

# ── Standings computation ──────────────────────────────────────────
def compute_standings(kerki_rounds, roster, mapper_sids):
    played = [r for r in kerki_rounds if r["last_lb"] and len(r["last_lb"]) > 0]

    # Map rotation: order maps appeared
    seen_maps = set()
    rotation = []
    for r in played:
        if r["map"] not in seen_maps:
            seen_maps.add(r["map"])
            rotation.append(r["map"])
    if len(rotation) != 3:
        print(f"  WARN: expected 3 unique maps, found {len(rotation)}: {rotation}", file=sys.stderr)

    # Warmup = first played round on each of the 3 unique maps
    warmup_keys = set()
    seen = set()
    for r in played:
        if r["map"] not in seen:
            seen.add(r["map"])
            warmup_keys.add((r["session"], r["num"]))
        if len(seen) == 3:
            break

    scoring_rounds = [r for r in played if (r["session"], r["num"]) not in warmup_keys]

    points = defaultdict(int)
    qualified_round = {}
    won_in_round = {}
    winners_in_order = []
    all_participants = set()
    rounds_played_by = defaultdict(int)
    finishes_by = defaultdict(int)
    mapper_points = defaultdict(int)

    for s_idx, r in enumerate(scoring_rounds, start=1):
        for sid in r["roster"]:
            rounds_played_by[sid] += 1
        lb = sorted(r["last_lb"], key=lambda x: x["pos"])
        for entry in lb:
            sid = entry["sid"]
            all_participants.add(sid)
            finishes_by[sid] += 1

        # ── Round winner (W determination) ─────────────────────────
        # Raw pos-1, but skip the round's own-map mapper if applicable.
        # NOTE: own-map-mapper-skip not implemented (would need per-map mapper config).
        # All nuisance/mapper wins block the round (no W produced).
        round_winner = lb[0]["sid"] if lb else None
        if (round_winner
            and round_winner not in mapper_sids
            and round_winner in qualified_round
            and qualified_round[round_winner] < s_idx
            and round_winner not in won_in_round):
            won_in_round[round_winner] = s_idx
            winners_in_order.append(round_winner)

        # ── Points (raw rank) ──────────────────────────────────────
        # Nuisance players' points go off-pool (mapper_points, uncapped).
        # Championship players accumulate at their RAW rank until 750.
        for rank0, entry in enumerate(lb):
            sid = entry["sid"]
            rank = rank0 + 1
            pts = points_for_pos(rank)
            if sid in mapper_sids:
                mapper_points[sid] += pts
                continue
            if sid in qualified_round:
                continue  # already capped at 750
            new_total = points[sid] + pts
            if new_total >= FINALIST_THRESHOLD:
                points[sid] = FINALIST_THRESHOLD
                qualified_round[sid] = s_idx
            else:
                points[sid] = new_total

    last_round_idx = len(scoring_rounds)
    finalists_valid = {sid: q for sid, q in qualified_round.items() if q < last_round_idx}
    winners_final = [sid for sid in winners_in_order if sid in finalists_valid][:5]
    winner_set = set(winners_final)

    def name_of(sid):
        return roster.get(sid, f"<{sid}>")

    qualifies = lambda sid: rounds_played_by[sid] >= MIN_ROUNDS_PLAYED and finishes_by[sid] >= 1

    # Build winners detail (winner is raw pos 1 of the won round)
    winners_out = []
    for i, sid in enumerate(winners_final, 1):
        won_idx = won_in_round[sid]
        won_round = scoring_rounds[won_idx - 1]
        lb = sorted(won_round["last_lb"], key=lambda x: x["pos"])
        winner_entry = next((e for e in lb if e["sid"] == sid), None)
        winners_out.append({
            "rank": i, "sid": sid, "name": name_of(sid),
            "won_scoring_round": won_idx,
            "qualified_scoring_round": qualified_round[sid],
            "map": won_round["map"],
            "time": winner_entry["time"] if winner_entry else "?",
        })

    finalists_only = sorted(
        (sid for sid in finalists_valid if sid not in winner_set),
        key=lambda s: finalists_valid[s],
    )
    finalists_out = [
        {"rank": i, "sid": sid, "name": name_of(sid),
         "qualified_scoring_round": finalists_valid[sid]}
        for i, sid in enumerate(finalists_only, 1)
    ]

    # Others
    rest = []
    excluded = []
    for sid in all_participants:
        if sid in winner_set or sid in finalists_valid:
            continue
        if sid in mapper_sids:
            continue
        if not qualifies(sid):
            excluded.append(sid)
            continue
        rest.append((sid, points[sid]))
    rest.sort(key=lambda x: (-x[1], name_of(x[0]).lower()))
    others_out = [
        {"rank": i, "sid": sid, "name": name_of(sid), "points": pts,
         "rounds_played": rounds_played_by[sid], "finishes": finishes_by[sid]}
        for i, (sid, pts) in enumerate(rest, 1)
    ]

    # Nuisance — mappers + playtesters listed if they participated
    nuisance_out = []
    for sid in sorted(mapper_sids, key=lambda s: -mapper_points[s]):
        if sid in all_participants or rounds_played_by[sid] > 0:
            nuisance_out.append({
                "sid": sid, "name": name_of(sid), "points": mapper_points[sid],
                "rounds_played": rounds_played_by[sid], "finishes": finishes_by[sid],
            })

    excluded_out = [
        {"sid": sid, "name": name_of(sid), "points": points[sid],
         "rounds_played": rounds_played_by[sid], "finishes": finishes_by[sid]}
        for sid in sorted(excluded, key=lambda s: -rounds_played_by[s])
    ]

    return {
        "rotation": rotation,
        "warmup_round_keys": sorted(warmup_keys),
        "scoring_rounds_count": len(scoring_rounds),
        "winners": winners_out,
        "finalists": finalists_out,
        "others": others_out,
        "nuisance": nuisance_out,
        "excluded": excluded_out,
    }

# ── xlsx writer ─────────────────────────────────────────────────────
def write_to_xlsx(kerki_id, standings, header_text, project_dir=PROJECT_DIR, dry_run=False):
    from openpyxl import load_workbook
    fname, sheet, idx = xlsx_target(kerki_id)
    col = OFFSETS[idx]
    name_col, points_col = col + 1, col + 2

    path = os.path.join(project_dir, fname)
    if not dry_run:
        backup = path + ".bak"
        shutil.copy(path, backup)
        print(f"  Backup: {backup}")

    wb = load_workbook(path)
    if sheet not in wb.sheetnames:
        wb.create_sheet(sheet)
    ws = wb[sheet]

    # Clear cols col..points_col rows 1-60
    for r in range(1, 60):
        for c in (col, name_col, points_col):
            ws.cell(row=r, column=c).value = None

    today = _date.today()
    date_str = f"{today.day}/{today.month}/{today.year % 100}"

    ws.cell(row=1, column=col).value = date_str
    ws.cell(row=2, column=col).value = header_text
    ws.cell(row=3, column=col).value = "Placement"
    ws.cell(row=3, column=name_col).value = "Name"
    ws.cell(row=3, column=points_col).value = "Points"

    row = 4
    ws.cell(row=row, column=col).value = "Winners"; row += 1
    for w in standings["winners"]:
        ws.cell(row=row, column=col).value = w["rank"]
        ws.cell(row=row, column=name_col).value = w["name"]
        row += 1

    if standings["finalists"]:
        ws.cell(row=row, column=col).value = "Finalists"; row += 1
        for f in standings["finalists"]:
            ws.cell(row=row, column=col).value = f["rank"]
            ws.cell(row=row, column=name_col).value = f["name"]
            row += 1

    if standings["others"]:
        ws.cell(row=row, column=col).value = "Other"; row += 1
        for o in standings["others"]:
            ws.cell(row=row, column=col).value = o["rank"]
            ws.cell(row=row, column=name_col).value = o["name"]
            ws.cell(row=row, column=points_col).value = o["points"]
            row += 1

    if standings["nuisance"]:
        ws.cell(row=row, column=col).value = "Nuisance"; row += 1
        for i, n in enumerate(standings["nuisance"], 1):
            ws.cell(row=row, column=col).value = i
            ws.cell(row=row, column=name_col).value = n["name"]
            ws.cell(row=row, column=points_col).value = n["points"]
            row += 1

    if not dry_run:
        wb.save(path)
        print(f"  Wrote {fname} sheet '{sheet}' col {col}")

# ── CLI ─────────────────────────────────────────────────────────────
def main():
    p = argparse.ArgumentParser()
    p.add_argument("kerki", type=int, help="Kerki number, e.g. 35")
    p.add_argument("--log", default=DEFAULT_LOG, help="Path to LiveLeaderboardLogger.log")
    p.add_argument("--date", default=None, help="Restrict to YYYY-MM-DD (default: any date)")
    p.add_argument("--mappers", default="",
                   help="Comma-separated steamIDs of championship-excluded players (mappers + playtesters)")
    p.add_argument("--header", default=None,
                   help="R2 cell content. Default: 'Kerki Comp #N - Maps by ?, ?, ?'")
    p.add_argument("--write-xlsx", action="store_true", help="Write to the xlsx (creates .bak first)")
    args = p.parse_args()

    mapper_sids = set(s.strip() for s in args.mappers.split(",") if s.strip())

    print(f"== Kerki #{args.kerki} ==")
    print(f"  log: {args.log}")
    print(f"  date filter: {args.date or '(any)'}")
    print(f"  mappers (exc.): {sorted(mapper_sids) or '(none)'}")

    kerki_rounds, roster = parse_log(args.log, args.kerki, restrict_date=args.date)
    if not kerki_rounds:
        print(f"  ERROR: no rounds tagged 'Kerki #{args.kerki}' found", file=sys.stderr)
        sys.exit(1)
    print(f"  kerki rounds parsed: {len(kerki_rounds)}")
    print(f"  roster size: {len(roster)}")

    standings = compute_standings(kerki_rounds, roster, mapper_sids)
    print(f"\n  rotation: {[m.replace(f'Kerki #{args.kerki} - ','') for m in standings['rotation']]}")
    print(f"  scoring rounds: {standings['scoring_rounds_count']}")

    print(f"\nWinners ({len(standings['winners'])}):")
    for w in standings["winners"]:
        m = w["map"].replace(f"Kerki #{args.kerki} - ", "")
        print(f"  {w['rank']}. {w['name']:25s}  S{w['won_scoring_round']} {m}  {w['time']}  (qualified S{w['qualified_scoring_round']})")

    print(f"\nFinalists ({len(standings['finalists'])}):")
    for f in standings["finalists"]:
        print(f"  {f['rank']}. {f['name']:25s}  qualified S{f['qualified_scoring_round']}")

    print(f"\nOthers ({len(standings['others'])}):")
    for o in standings["others"]:
        print(f"  {o['rank']:2d}. {o['name']:25s} {o['points']:>4} pts  ({o['rounds_played']}r/{o['finishes']}f)")

    print(f"\nNuisance ({len(standings['nuisance'])}):")
    for n in standings["nuisance"]:
        print(f"   - {n['name']:25s} {n['points']} pts  ({n['rounds_played']}r/{n['finishes']}f)")

    if standings["excluded"]:
        print(f"\nExcluded — <{MIN_ROUNDS_PLAYED}r or 0f ({len(standings['excluded'])}):")
        for e in standings["excluded"]:
            print(f"   - {e['name']:25s} pts={e['points']:>4}  rounds={e['rounds_played']}  finishes={e['finishes']}")

    out_path = os.path.join(PROJECT_DIR, f"kerki_{args.kerki}.json")
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump({"kerki": args.kerki, "header": args.header, **standings}, f, indent=2, ensure_ascii=False)
    print(f"\nJSON: {out_path}")

    if args.write_xlsx:
        header = args.header or f"Kerki Comp #{args.kerki} - Maps by ?, ?, ?"
        print(f"\nWriting xlsx (header='{header}'):")
        write_to_xlsx(args.kerki, standings, header)
        print(f"  Done. Now run: python build_kerki.py")

if __name__ == "__main__":
    main()
