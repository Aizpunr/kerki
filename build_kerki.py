"""Build kerki.json from the 3 Kerki Comp Results xlsx files."""
import json, re, os
from collections import defaultdict
from openpyxl import load_workbook

os.chdir(os.path.dirname(os.path.abspath(__file__)))

# ── Name normalization ──────────────────────────────────────────────
def strip_tag(name):
    return re.sub(r'\[.*?\]\s*', '', name).strip()

# Kerki-specific canonical map (overrides COTD where needed)
CANONICAL = {
    'rtube': ['rtyyyyb', 'rtube', '[TBD] rtube', '[TBD]rtube', '[BAP]rtube',
              '[Toob]rtube', '[dumb]rtube', '[sad]rtube', '[SUCK] rtube',
              '[TOB] rtyyyyb', '[ZET] rtyyyyb', '[ZET]rtyyyyb', '[TBD] rtyyyyb',
              'rtubert the farmer'],
    'Quickracer10': ['quickracer10', 'quickracer', '[KURK] Quickracer10',
                     '[AJSE] Quickracer10', '[ASJE] Quickracer10', '[CC] Quickracer10',
                     'Quick'],
    'Warsnac': ['Warcans', '[CHR]Warcans', '[BAP]Warsnac', '[CHR]Warsnac', '[old]Warsnac'],
    'JakeAdjacent': ['SadD0ge', '[CD]SadD0ge', '[SWMG]SadD0ge', '[SWMG]JakeAdjacent'],
    'Naomi': ['Nyaomi', '[meow]Nyaomi', '[POIN]Fwogiie', '[RTR]Fwogiie',
              '[frog]Fwogiie', 'Fwogiie', '[ASJE]Naomi'],
    'Tommygaming': ['Tommygaming5132', 'TommyGaming5132', '[CSC]Tommygaming',
                    '[CSC]Tommygaming5132', '[TG]Tommygaming5132',
                    '[OOPS]Tommygaming', '[jofk]Tommygaming',
                    '[CSC]Tommygaming6132', '[C3PO]Tommygaming5132'],
    'An Actual g00se': ['g00se', '[CSC] An Actual g00se', '[Err] An Actual g00se',
                        '[CSC] BaBa is g00se', '[CSC] CantFindTheg00se'],
    'St Nicholas': ['St Nic', 'St. Nicholas', 'St NIcholas', '[cozy]St Nicholas'],
    'JobW': ['Job'],
    'AndMe': ['AndMe16', 'AndMe17', 'Andme17', '[COMY]AndMe17', '[CSC]AndMe17',
              '[ORIG]AndMe16', 'AndMe14', 'AndMe15'],
    'Hydro': ['[BFP] Hydro', '[CTR]Hydro', '[RTR] Hydro', '[SLOW] Hydro',
              '[WOW]Hydro', '[ZOMN] Hydro'],
    'ZOMAN': ['[ARMS] ZOMAN', '[Bath] ZOMAN', '[Blub] ZOMAN', '[Choo] ZOMAN',
              '[KUNG] ZOMAN', '[SLOW]ZOMAN', '[TOOB] ZOMAN'],
    'Sandals': ['[CTR] Sandals'],
    'Kernkob': ['kernkob', '[CTR]Kernkob', 'kernbooper'],
    'RoundNzt': ['RoundNzt'],
    'Lexer': ['[BRIT] Lexer'],
    'PandaMane': ['[FOV]PandaMane', '[FPV]PandaMane', '[CHEZ]PandaMane'],
    'Linzi': [],
    'Sterben': ['[BAP]Sterben', '[PNCK]Mini P.E.K.K.A', '[FPV]Lλmbda', '[PCDJ]Sterben'],
    'Lazy_Echidna': ['Lazy Echidna', '[NIL]Lazy_Echidna', '[TOG]Lazy_Echidna'],
    'Metalted': ['[ZMS] Metalted', '[ZMS]Metalted'],
    'LILWOOLEY': ['[ZET]LILWOOLEY', '[ZST]LILWOOLEY'],
    'Mortishade': ['[CTR]Mortishade', '[bam]Mortishade'],
    'Shadynook': ['[CSC] Shadynook', '[CSC]Shadynook', '[LATE]Shadynook'],
    'MackCheesy': ['[CHEZ]MackCheesy', '[ZET]MackCheesy'],
    'Murrl': ['[BAP]Murrl', '[Burp]Murrl', '[Toob]Murrl'],
    'readfreak7': ['[HRR] readfreak7', '[HRR]readfreak7', '[PFE] readfreak7'],
    'redal': ['[CSC] redal'],
    'Lamp': ['[CTR]Lamp', '[The]Lamp', '[bam]Lamp'],
    'schmxrg': ['[6dog] schmxrg', '[dogg] schmxrg', '[goat] schmxrg'],
    'XpERt': ['[TBD]XpERt'],
    'Jakie': ['[CD] Jakie', '[ZET] Jakie'],
    'R0nanC': ['[CTR] R0nanC', '[CTR]R0nanC', 'R0nanc'],
    'Northwind': [],
    'jandje': ['[BFP] jandje', '[CTR] jandje', '[CTR]jandje'],
    'Kaiser64': ['[TEA]Kaiser64'],
    'microways': ['[Quac] microways'],
    'ping': ['[bad] ping', '[boom] ping', '[no] ping', '[pong]ping'],
    'agix': ['[GYMC] agix'],
    'DeeDeeNaNaNa': ['[CSC] DeeDeeNaNaNa', '[CSC]DeeDeeNaNaNa'],
    'Redstony': ['[Stc3]Redstony', '[TILT]Redstony'],
    'RadAbsRad': ['[Meow]RadAbsRad'],
    'Hi Im Yolo': ['[HRR]Hi Im Yolo', '[RIP]Hi Im Yolo'],
    'variableferret': ['[CSC] variableferret'],
    'Noxitu': [],
    'vortex': ['[mib]vortex'],
    'brrryy': ['brryyy'],
    'Orthros': [],
    'MALIYO': [],
    'dlwldms': [],
    'SHOOOOOOP': [],
    'Jinx': ['[DCS] Jinx'],
    'Dartho Mas': [],
    'Kodiak': [],
    'DeiRex': [],
    'palfly': [],
    'Matic_D': [],
    'maxie12': [],
    'M4rv': [],
    'Catch': [],
    'joshuadwyer79': [],
    'Heart-TGV': ['[TTR]Heart-TGV'],
    'Hawk': [],
    'graysonvitek88': [],
}

NAME_MAP = {}
for canonical, aliases in CANONICAL.items():
    for alias in aliases:
        NAME_MAP[alias] = canonical

def normalize(name):
    if not name:
        return None
    name = str(name).strip()
    if not name:
        return None
    # Direct lookup
    if name in NAME_MAP:
        return NAME_MAP[name]
    # Already canonical
    if name in CANONICAL:
        return name
    # Strip tag and retry
    stripped = strip_tag(name)
    if stripped in NAME_MAP:
        return NAME_MAP[stripped]
    if stripped in CANONICAL:
        return stripped
    # Case-insensitive match
    lower = stripped.lower()
    for canon in CANONICAL:
        if canon.lower() == lower:
            return canon
    return stripped

# ── Troll kerkis ────────────────────────────────────────────────────
TROLL_KERKIS = {10, 20, 30}

# ── Read xlsx data ──────────────────────────────────────────────────
def read_kerki_tab(ws, tab_start_kerki, col_offsets=None):
    """Read kerkis from a worksheet tab. Each kerki uses 6 columns (or 5 data cols + 1 blank)."""
    kerkis = []
    if col_offsets is None:
        col_offsets = [1, 7, 13, 19, 25]  # standard 5 kerkis per tab

    for idx, col_start in enumerate(col_offsets):
        kerki_id = tab_start_kerki + idx

        # Check if there's data (date in row 1 OR header in row 2)
        date_val = ws.cell(row=1, column=col_start).value
        header_val = ws.cell(row=2, column=col_start).value
        if date_val is None and header_val is None:
            continue

        winners = []
        finalists = []
        others = []
        current_section = None

        for row in range(3, ws.max_row + 1):
            placement = ws.cell(row=row, column=col_start).value
            name = ws.cell(row=row, column=col_start + 1).value

            if placement is None and name is None:
                continue

            # Detect section headers
            p_str = str(placement).strip() if placement else ''
            if p_str.lower() in ('winners', 'winner'):
                current_section = 'winners'
                continue
            elif p_str.lower() in ('finalists', 'finalist'):
                current_section = 'finalists'
                continue
            elif p_str.lower() in ('other', 'others'):
                current_section = 'others'
                continue
            elif p_str.lower() == 'placement':
                continue

            if name is None:
                continue

            name = normalize(name)
            if name is None:
                continue

            # Skip non-player text entries
            if any(x in str(name).lower() for x in ['data lost', 'tech issues', 'did not reach']):
                continue

            # Skip known nuisances/spectators
            if name.lower() in ('justmaki', 'kernkob') and current_section == 'others':
                # They might be nuisances in others, but keep them if they're winners/finalists
                pass  # actually keep them — they might genuinely be "others" in some kerkis

            if current_section == 'winners':
                winners.append(name)
            elif current_section == 'finalists':
                finalists.append(name)
            elif current_section == 'others':
                others.append(name)

        if winners or finalists:
            lobby = len(set(winners + finalists + others))
            kerkis.append({
                'id': kerki_id,
                'date': str(date_val) if date_val else '?',
                'troll': kerki_id in TROLL_KERKIS,
                'winners': winners,
                'finalists': finalists,
                'others': others,
                'lobby_size': lobby,
            })

    return kerkis

all_kerkis = []

# File 1: Kerkis 1-15
wb1 = load_workbook('Kerki Comp Results.xlsx', read_only=True, data_only=True)
for tab_name, start_id in [('Kerki 1-5', 1), ('Kerki 6-10', 6), ('Kerki 11-15', 11)]:
    if tab_name in wb1.sheetnames:
        all_kerkis.extend(read_kerki_tab(wb1[tab_name], start_id))
wb1.close()

# File 2: Kerkis 16-30
wb2 = load_workbook('Kerki Comp Results 16-30.xlsx', read_only=True, data_only=True)
for tab_name, start_id in [('Kerki 16-20', 16), ('Kerki 21-25', 21), ('Kerki 26-30', 26)]:
    if tab_name in wb2.sheetnames:
        all_kerkis.extend(read_kerki_tab(wb2[tab_name], start_id))
wb2.close()

# File 3: Kerkis 31+
wb3 = load_workbook('Kerki Comp Results 31+.xlsx', read_only=True, data_only=True)
for tab_name, start_id in [('Kerki 31-35', 31),]:
    if tab_name in wb3.sheetnames:
        all_kerkis.extend(read_kerki_tab(wb3[tab_name], start_id))
wb3.close()

# Sort by id
all_kerkis.sort(key=lambda k: k['id'])

print(f"Loaded {len(all_kerkis)} kerkis")
for k in all_kerkis:
    t = " [TROLL]" if k['troll'] else ""
    print(f"  Kerki #{k['id']}: {len(k['winners'])}W + {len(k['finalists'])}F + {len(k['others'])}O = {k['lobby_size']} players{t}")

# ── Build player stats ──────────────────────────────────────────────
players = defaultdict(lambda: {
    'apps': 0, 'wins': 0,
    'w1': 0, 'w2': 0, 'w3': 0, 'w4': 0, 'w5': 0,
    'finalist': 0, 'best': 99,
    'history': [],
})

for k in all_kerkis:
    all_participants = set()

    for i, name in enumerate(k['winners']):
        placement = i + 1
        p = players[name]
        p['apps'] += 1
        p['wins'] += 1
        p[f'w{placement}'] += 1
        if placement < p['best']:
            p['best'] = placement
        p['history'].append({'k': k['id'], 'result': f'w{placement}'})
        all_participants.add(name)

    for name in k['finalists']:
        if name in all_participants:
            continue  # shouldn't happen but safety
        p = players[name]
        p['apps'] += 1
        p['finalist'] += 1
        if 6 < p['best']:  # finalist = rank 6 conceptually
            p['best'] = 6
        p['history'].append({'k': k['id'], 'result': 'f'})
        all_participants.add(name)

    for name in k['others']:
        if name in all_participants:
            continue
        p = players[name]
        p['apps'] += 1
        p['history'].append({'k': k['id'], 'result': 'o'})
        all_participants.add(name)

# Build output
player_list = []
for name, p in sorted(players.items(), key=lambda x: (-x[1]['wins'], -x[1]['apps'])):
    player_list.append({
        'name': name,
        'apps': p['apps'],
        'wins': p['wins'],
        'w1': p['w1'],
        'w2': p['w2'],
        'w3': p['w3'],
        'w4': p['w4'],
        'w5': p['w5'],
        'finalist': p['finalist'],
        'best': p['best'] if p['best'] < 99 else None,
        'history': p['history'],
    })

# ── ATP-style ranking ──────────────────────────────────────────────
RANK_POINTS = {'w1': 200, 'w2': 140, 'w3': 100, 'w4': 70, 'w5': 40, 'f': 10}
RANK_WINDOW = 26
RANK_BEST_OF = 18

non_troll_ids = sorted(k['id'] for k in all_kerkis if not k['troll'])
eligible = set(non_troll_ids[-RANK_WINDOW:])

ranking_list = []
for pl in player_list:
    # Filter history to eligible window — only scoring results (exclude 'o')
    scoring_hist = []
    window_apps = 0
    for h in pl['history']:
        if h['k'] not in eligible:
            continue
        window_apps += 1
        pts = RANK_POINTS.get(h['result'], 0)
        if pts > 0:
            scoring_hist.append({'k': h['k'], 'result': h['result'], 'points': pts})

    if not scoring_hist:
        continue

    # Sort by points desc, take best RANK_BEST_OF
    scoring_hist.sort(key=lambda x: -x['points'])
    for i, e in enumerate(scoring_hist):
        e['counted'] = i < RANK_BEST_OF

    total = sum(e['points'] for e in scoring_hist if e['counted'])
    total_all = sum(e['points'] for e in scoring_hist)
    counted = sum(1 for e in scoring_hist if e['counted'])
    dropped = len(scoring_hist) - counted

    w_results = [e['result'] for e in scoring_hist]
    ranking_list.append({
        'name': pl['name'],
        'points': total,
        'points_all': total_all,
        'apps': window_apps,
        'wins': sum(1 for r in w_results if r.startswith('w')),
        'w1': w_results.count('w1'), 'w2': w_results.count('w2'),
        'w3': w_results.count('w3'), 'w4': w_results.count('w4'),
        'w5': w_results.count('w5'), 'finalist': w_results.count('f'),
        'counted': counted, 'dropped': dropped,
        'history': sorted(scoring_hist, key=lambda x: x['k']),
    })

ranking_list.sort(key=lambda x: (-x['points'], -x['wins'], -x['w1']))
for i, r in enumerate(ranking_list):
    r['rank'] = i + 1

print(f"\nATP Ranking (window: #{min(eligible)}-#{max(eligible)}, best {RANK_BEST_OF} of {len(eligible)}):")
for r in ranking_list[:15]:
    print(f"  #{r['rank']:>2}  {r['name']:20s}  {r['points']:>5} pts  "
          f"({r['apps']} apps, {r['counted']} counted, {r['dropped']} dropped)")

# ── Glicko-style skill rating ──────────────────────────────────────
import math

GLICKO_PERF = {'w1': 2100, 'w2': 1950, 'w3': 1875, 'w4': 1850, 'w5': 1800, 'f': 1700, 'o': 1500}
GLICKO_TAU = 200          # system constant (convergence speed)
GLICKO_MU_START = 1500    # starting rating
GLICKO_SIGMA_START = 350  # starting uncertainty
GLICKO_SIGMA_FLOOR = 80   # min uncertainty (always some movement)
GLICKO_SIGMA_CAP = 350    # max uncertainty
GLICKO_SIGMA_DECAY = 15   # sigma growth per missed kerki
GLICKO_MU_DECAY = 0.02    # 2% rating decay toward 1500 per missed kerki
GLICKO_SURPRISE_THRESHOLD = 200  # |perf - mu| above this widens sigma
GLICKO_SURPRISE_FACTOR = 0.3     # how much sigma widens on surprise

troll_ids = set(k['id'] for k in all_kerkis if k['troll'])
glicko_non_troll = sorted(k['id'] for k in all_kerkis if not k['troll'])

# Build per-player event list
glicko_events = {}
for pl in player_list:
    for h in pl['history']:
        if h['k'] in troll_ids:
            continue
        glicko_events.setdefault(pl['name'], []).append((h['k'], h['result']))
for name in glicko_events:
    glicko_events[name].sort()

glicko_ratings = {}

for kid in glicko_non_troll:
    # Who played this kerki
    participants = set()
    for name, events in glicko_events.items():
        for k, result in events:
            if k == kid:
                participants.add(name)
                break

    # Decay + sigma growth for non-participants
    for name in glicko_ratings:
        if name not in participants:
            r = glicko_ratings[name]
            # Permanent decay toward 1500
            r['mu'] = GLICKO_MU_START + (r['mu'] - GLICKO_MU_START) * (1 - GLICKO_MU_DECAY)
            # Sigma grows (less certain)
            r['sigma'] = min(r['sigma'] + GLICKO_SIGMA_DECAY, GLICKO_SIGMA_CAP)

    # Update participants
    for name, events in glicko_events.items():
        for k, result in events:
            if k != kid:
                continue
            if name not in glicko_ratings:
                glicko_ratings[name] = {'mu': GLICKO_MU_START, 'sigma': GLICKO_SIGMA_START, 'history': []}
            r = glicko_ratings[name]
            perf = GLICKO_PERF.get(result, 1500)

            # Surprise detection: widen sigma if result far from expected
            surprise = abs(perf - r['mu'])
            if surprise > GLICKO_SURPRISE_THRESHOLD:
                r['sigma'] = min(r['sigma'] + (surprise - GLICKO_SURPRISE_THRESHOLD) * GLICKO_SURPRISE_FACTOR,
                                 GLICKO_SIGMA_CAP)

            # Update rating
            lr = r['sigma']**2 / (r['sigma']**2 + GLICKO_TAU**2)
            r['mu'] = r['mu'] + lr * (perf - r['mu'])
            r['sigma'] = max(math.sqrt((1 - lr) * r['sigma']**2), GLICKO_SIGMA_FLOOR)
            r['history'].append({'k': kid, 'mu': round(r['mu'], 1), 'sigma': round(r['sigma'], 1), 'result': result})
            break

# Build glicko output — only players with scoring results
glicko_list = []
for name, r in glicko_ratings.items():
    if not any(h['result'] != 'o' for h in r['history']):
        continue
    apps = len(r['history'])
    results = [h['result'] for h in r['history']]
    wins = sum(1 for x in results if x.startswith('w'))
    glicko_list.append({
        'name': name,
        'mu': round(r['mu'], 1),
        'sigma': round(r['sigma'], 1),
        'apps': apps,
        'wins': wins,
        'w1': results.count('w1'), 'w2': results.count('w2'),
        'w3': results.count('w3'), 'w4': results.count('w4'),
        'w5': results.count('w5'), 'finalist': results.count('f'),
        'history': r['history'],
    })

glicko_list.sort(key=lambda x: -x['mu'])
for i, g in enumerate(glicko_list):
    g['rank'] = i + 1

print(f"\nGlicko Skill Rating (decay={GLICKO_MU_DECAY}, surprise={GLICKO_SURPRISE_THRESHOLD}):")
for g in glicko_list[:15]:
    print(f"  #{g['rank']:>2}  {g['name']:20s}  {g['mu']:>6.0f} +/-{g['sigma']:<.0f}  ({g['apps']} apps)")

output = {
    'meta': {
        'total_kerkis': len(all_kerkis),
        'total_players': len(player_list),
        'last_kerki': max(k['id'] for k in all_kerkis) if all_kerkis else 0,
    },
    'kerkis': all_kerkis,
    'players': player_list,
    'ranking': {
        'window': RANK_WINDOW,
        'best_of': RANK_BEST_OF,
        'window_start': min(eligible),
        'window_end': max(eligible),
        'eligible_kerkis': len(eligible),
        'players': ranking_list,
    },
    'glicko': {
        'start': GLICKO_MU_START,
        'decay': GLICKO_MU_DECAY,
        'players': glicko_list,
    },
}

with open('kerki.json', 'w', encoding='utf-8') as f:
    json.dump(output, f, ensure_ascii=False, indent=1)

print(f"\nOutput: kerki.json — {len(player_list)} players across {len(all_kerkis)} kerkis")
print(f"\nTop 10 by wins:")
for p in player_list[:10]:
    pct = round((p['wins'] + p['finalist']) / p['apps'] * 100) if p['apps'] else 0
    print(f"  {p['name']:20s}  Apps:{p['apps']:3d}  Wins:{p['wins']:2d}  "
          f"(1st:{p['w1']} 2nd:{p['w2']} 3rd:{p['w3']})  Fin:{p['finalist']:2d}  Win%:{pct}%")
